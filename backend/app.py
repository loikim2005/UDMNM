"""
Flask API: train / predict / attendance / students — kết nối UI + MySQL.
"""
from __future__ import annotations

import json
import logging
import os
import re
import shutil
from pathlib import Path
import sys
import uuid
import io
import math
import traceback
from datetime import datetime
from decimal import Decimal
from typing import Any, Dict, List, Optional, Set, Tuple

import cv2
import numpy as np
import pymysql
import tensorflow as tf
import time
from dotenv import load_dotenv
from flask import Flask, g, jsonify, request, send_file
from flask_cors import CORS
from itsdangerous import BadSignature, SignatureExpired, URLSafeTimedSerializer
from werkzeug.security import check_password_hash, generate_password_hash
from werkzeug.exceptions import HTTPException
from werkzeug.wrappers import Response as WzResponse
from email_service import send_email, send_email_async

# face_duplicate vẫn được dùng cho đăng ký (file tồn tại), nhưng đăng ký hiện chặn bằng embeddings.npz
from train import default_train_epochs, train_from_directory
from utils import (
    data_dir,
    decode_base64_to_numpy_rgb,
)
from face_embedder import embedding_from_face_rgb, embedding_from_image_rgb
from yolo_detector import get_yolo_detector, get_yolo_status
from recognition import (
    load_embeddings_db as load_recognition_embeddings_db,
    recognize_faces_in_frame,
    reset_embeddings_cache,
)

# Nạp biến môi trường từ .env (SMTP, MySQL, ...)
load_dotenv()

app = Flask(__name__)
app.config["SECRET_KEY"] = os.environ.get("FLASK_SECRET_KEY", "face-attendance-dev-key-change-me")
CORS(
    app,
    resources={r"/*": {"origins": os.environ.get("CORS_ORIGINS", "*").split(",")}},
    # Dùng Bearer token qua header, không dùng cookie -> không cần credentials.
    # Tránh lỗi browser khi origins="*" nhưng lại bật credentials.
    supports_credentials=False,
    allow_headers=["Content-Type", "Authorization"],
)

_FACE_SIZE = (160, 160)
_CASCADE_PATH = os.path.join(cv2.data.haarcascades, "haarcascade_frontalface_default.xml")
_haar: Optional[cv2.CascadeClassifier] = None
_embed_model: Optional[tf.keras.Model] = None
_register_dup_cache: Dict[str, Any] = {"students": {}}
_attendance_session_seen: Set[str] = set()
_attendance_session_day: str = ""


def _configure_logging_once() -> logging.Logger:
    """
    Cấu hình logger cho backend.
    - Dùng logging (không dùng print)
    - Flush realtime (StreamHandler)
    - Format rõ ràng cho CMD
    """
    logger = logging.getLogger("mnm")
    if logger.handlers:
        return logger

    level = os.environ.get("LOG_LEVEL", "INFO").strip().upper() or "INFO"
    logger.setLevel(getattr(logging, level, logging.INFO))

    handler = logging.StreamHandler(stream=sys.stdout)
    fmt = logging.Formatter(
        fmt="[%(asctime)s] %(levelname)s %(message)s",
        datefmt="%Y-%m-%d %H:%M:%S",
    )
    handler.setFormatter(fmt)
    logger.addHandler(handler)
    logger.propagate = False
    return logger


_log = _configure_logging_once()


def _should_log_requests() -> bool:
    return os.environ.get("LOG_REQUESTS", "1").strip().lower() not in ("0", "false", "no")


def _client_ip() -> str:
    # Nếu chạy sau proxy thì lấy X-Forwarded-For (phần đầu)
    xff = (request.headers.get("X-Forwarded-For") or "").strip()
    if xff:
        return xff.split(",")[0].strip()
    return request.remote_addr or "-"


def _safe_request_data_for_log() -> Dict[str, Any]:
    """
    Log request data khi lỗi:
    - JSON body nếu có
    - form nếu có
    - query string
    Tránh crash logger nếu body không đọc được.
    """
    out: Dict[str, Any] = {}
    try:
        js = request.get_json(silent=True)
    except Exception:
        js = None
    if js is not None:
        out["json"] = js
    try:
        if request.form:
            out["form"] = {k: request.form.get(k) for k in request.form.keys()}
    except Exception as e:
        _log.error(f"[ATTENDANCE] Save failed: {e}")
    try:
        if request.args:
            out["args"] = {k: request.args.get(k) for k in request.args.keys()}
    except Exception:
        pass
    try:
        if request.files:
            out["files"] = list(request.files.keys())
    except Exception:
        pass
    return out


def _json_dumps_compact(v: Any) -> str:
    try:
        return json.dumps(v, ensure_ascii=False, separators=(",", ":"), default=str)
    except Exception:
        return str(v)


def _log_login(success: bool, username: str, reason: Optional[str] = None) -> None:
    u = (username or "").strip() or "-"
    if success:
        _log.info(f"[LOGIN SUCCESS] user={u}")
    else:
        if reason:
            _log.warning(f"[LOGIN FAILED] user={u} reason={reason}")
        else:
            _log.warning(f"[LOGIN FAILED] user={u}")


def _log_recognition(student_code: Optional[str], confidence: float, success: bool) -> None:
    conf = float(confidence)
    pct = round(conf * 100.0, 2)
    if success and student_code:
        _log.info(f"[RECOGNITION] MSSV={student_code} | Confidence={conf:.4f} ({pct}%) | SUCCESS")
    else:
        _log.info(f"[RECOGNITION] UNKNOWN | Confidence={conf:.4f} ({pct}%) | FAILED")


def _log_register(student_code: str, images: int, status: str) -> None:
    sc = (student_code or "").strip() or "-"
    _log.info(f"[REGISTER] MSSV={sc} | Images={int(images)} | {status}")


def _resolve_attendance_identity(student_db_id: Any, student_code: Any) -> Tuple[Optional[int], Optional[str]]:
    sid: Optional[int] = None
    scode = (str(student_code).strip() if student_code is not None else "") or None
    try:
        if student_db_id is not None and str(student_db_id).strip() != "":
            sid = int(student_db_id)
    except Exception:
        sid = None
    if sid is not None:
        return sid, scode
    if scode:
        row = fetch_student_by_code(scode)
        if row and row.get("id") is not None:
            try:
                sid = int(row["id"])
            except Exception:
                sid = None
    return sid, scode


def _attendance_exists_for_student_today(cur, student_db_id: Optional[int], student_code: Optional[str]) -> bool:
    if student_db_id is None and not student_code:
        return False
    cur.execute(
        "SELECT 1 FROM attendance "
        "WHERE DATE(checked_at)=CURDATE() "
        "AND ("
        "(%s IS NOT NULL AND student_db_id=%s) "
        "OR (%s IS NOT NULL AND student_code=%s)"
        ") "
        "LIMIT 1",
        (student_db_id, student_db_id, student_code, student_code),
    )
    return cur.fetchone() is not None


def _attendance_exists_for_student_subject_today(
    cur,
    student_db_id: Optional[int],
    student_code: Optional[str],
    subject_id: Optional[int],
) -> bool:
    """
    Check duplicate theo ngày + môn (nếu có subject_id).
    """
    if student_db_id is None and not student_code:
        return False
    if subject_id is None:
        return _attendance_exists_for_student_today(cur, student_db_id, student_code)
    cur.execute(
        "SELECT 1 FROM attendance "
        "WHERE DATE(checked_at)=CURDATE() "
        "AND subject_id=%s "
        "AND ("
        "(%s IS NOT NULL AND student_db_id=%s) "
        "OR (%s IS NOT NULL AND student_code=%s)"
        ") "
        "LIMIT 1",
        (subject_id, student_db_id, student_db_id, student_code, student_code),
    )
    return cur.fetchone() is not None


def _session_seen_today(student_code: Optional[str]) -> bool:
    global _attendance_session_day, _attendance_session_seen
    if not student_code:
        return False
    today = datetime.now().strftime("%Y-%m-%d")
    if _attendance_session_day != today:
        _attendance_session_day = today
        _attendance_session_seen = set()
    return str(student_code) in _attendance_session_seen


def _mark_session_seen_today(student_code: Optional[str]) -> None:
    global _attendance_session_day, _attendance_session_seen
    if not student_code:
        return
    today = datetime.now().strftime("%Y-%m-%d")
    if _attendance_session_day != today:
        _attendance_session_day = today
        _attendance_session_seen = set()
    _attendance_session_seen.add(str(student_code))


def _moire_peakiness_score_from_face_bgr(face_bgr: np.ndarray) -> float:
    """
    Passive liveness (heuristic): FFT spectrum peakiness để bắt Moire pattern
    (đỉnh tần số bất thường do chụp qua màn hình/ảnh in).

    Trả về score (càng cao càng nghi ngờ). Ngưỡng so sánh lấy từ env:
    MOIRE_PEAKINESS_THRESHOLD (float). Nếu threshold <= 0 thì coi như tắt check.
    """
    if face_bgr is None or getattr(face_bgr, "size", 0) == 0:
        return 0.0
    if face_bgr.ndim != 3 or face_bgr.shape[2] != 3:
        raise ValueError("face_bgr phải có shape (H,W,3)")

    # Chuẩn hóa input để FFT ổn định hơn
    gray = cv2.cvtColor(face_bgr, cv2.COLOR_BGR2GRAY)
    # Giảm nhiễu trước FFT để tránh false positive (camera noise/aliasing)
    gray = cv2.GaussianBlur(gray, (5, 5), 0)
    gray = cv2.resize(gray, (256, 256), interpolation=cv2.INTER_AREA)
    x = gray.astype(np.float32) / 255.0
    x = x - float(np.mean(x))

    # Giảm artefact biên khi FFT
    win = np.hanning(x.shape[0]).astype(np.float32)
    x = x * (win[:, None] * win[None, :])

    # FFT2 -> log magnitude spectrum
    f = np.fft.fft2(x)
    fshift = np.fft.fftshift(f)
    mag = np.abs(fshift).astype(np.float32)
    mag = np.log1p(mag)

    h, w = mag.shape
    cy, cx = h // 2, w // 2
    yy, xx = np.ogrid[:h, :w]
    r = np.sqrt((yy - cy) ** 2 + (xx - cx) ** 2)

    # Bỏ vùng tần số thấp (trung tâm), chỉ xem phần high-frequency
    r0 = 0.10 * min(h, w)  # low-freq radius
    r1 = 0.45 * min(h, w)  # tránh sát biên (nhiễu resize/aliasing)
    mask = (r >= r0) & (r <= r1)
    vals = mag[mask]
    if vals.size < 64:
        return 0.0

    # Peakiness (robust): chuẩn hóa theo MAD để giảm phụ thuộc độ sáng/camera.
    med = float(np.median(vals))
    mad = float(np.median(np.abs(vals - med))) + 1e-6
    p995 = float(np.percentile(vals, 99.5))
    score = (p995 - med) / mad
    if not np.isfinite(score):
        return 0.0
    return float(score)


def _passive_liveness_check_moire(face_rgb: np.ndarray) -> Tuple[bool, float, float]:
    """
    Return (is_fraud, score, threshold).
    """
    th = float(os.environ.get("MOIRE_PEAKINESS_THRESHOLD", "0"))
    if th <= 0:
        return False, 0.0, th
    face_bgr = face_rgb[:, :, ::-1].copy()
    score = _moire_peakiness_score_from_face_bgr(face_bgr)
    return bool(score >= th), float(score), th


def _read_frame_rgb_from_request_json_or_multipart(body: Dict[str, Any]) -> Optional[np.ndarray]:
    """
    Ưu tiên multipart file `image`, fallback JSON base64 `image` hoặc `images[0]`.
    Trả về frame RGB uint8 hoặc None nếu không có ảnh.
    """
    frame_rgb: Optional[np.ndarray] = None
    if "image" in request.files:
        try:
            raw = request.files["image"].read()
            arr = np.frombuffer(raw, dtype=np.uint8)
            bgr = cv2.imdecode(arr, cv2.IMREAD_COLOR)
            if bgr is not None and bgr.size > 0:
                frame_rgb = bgr[:, :, ::-1].copy()
        except Exception:
            frame_rgb = None
    if frame_rgb is not None:
        return frame_rgb

    img_b64 = body.get("image")
    imgs = body.get("images") or []
    if not img_b64 and isinstance(imgs, list) and imgs:
        img_b64 = imgs[0]
    if not img_b64:
        return None
    try:
        return decode_base64_to_numpy_rgb(str(img_b64))
    except Exception:
        return None


def _recognize_and_attend_from_frame_rgb(
    *,
    frame_rgb: np.ndarray,
    latitude: Optional[float] = None,
    longitude: Optional[float] = None,
    subject_id: Optional[int] = None,
) -> Any:
    """
    YOLO detect nhiều khuôn mặt -> tạo embedding từng mặt -> match DB -> lưu điểm danh từng sinh viên hợp lệ.
    """
    if frame_rgb is None or frame_rgb.size == 0:
        return jsonify({"ok": False, "error": "Frame rỗng."}), 400

    frame_bgr = frame_rgb[:, :, ::-1].copy()
    max_faces = int(os.environ.get("MAX_MULTI_FACES", "3"))
    yolo_model_name = os.path.basename(os.environ.get("YOLO_MODEL_PATH", "yolov8n-face.pt").strip() or "yolov8n-face.pt")
    try:
        detector = get_yolo_detector()
        detections = detector.detect(frame_bgr, max_faces=max_faces)
    except FileNotFoundError:
        return (
            jsonify(
                {
                    "ok": False,
                    "error": "YOLO not ready",
                    "hint": f"Please download {yolo_model_name}",
                }
            ),
            200,
        )
    except Exception:
        yolo_info = get_yolo_status()
        if yolo_info.get("status") in ("missing_model", "error"):
            return (
                jsonify(
                    {
                        "ok": False,
                        "error": "YOLO not ready",
                        "hint": f"Please download {yolo_model_name}",
                    }
                ),
                200,
            )
        return jsonify({"ok": False, "error": "YOLO not ready"}), 200

    if not detections:
        _log_recognition(None, 0.0, False)
        return (
            jsonify(
                {
                    "ok": True,
                    "success": False,
                    "face_detected": False,
                    "matches": [],
                    "count": 0,
                    "message": "Không tìm thấy khuôn mặt",
                }
            ),
            200,
        )

    try:
        db = load_recognition_embeddings_db()
        matches = recognize_faces_in_frame(
            frame_rgb=frame_rgb,
            detections=detections,
            embeddings_db=db["embeddings"],
            labels=db["labels"],
            names=db["names"],
            fetch_student_by_code=fetch_student_by_code,
            max_faces=max_faces,
        )
    except (FileNotFoundError, ValueError) as e:
        return jsonify({"ok": False, "error": str(e)}), 400
    except Exception as e:
        return jsonify({"ok": False, "error": f"Nhận diện lỗi: {e}"}), 500

    if not matches:
        return (
            jsonify(
                {
                    "ok": True,
                    "success": False,
                    "face_detected": True,
                    "matches": [],
                    "count": 0,
                    "message": "Không có sinh viên hợp lệ trong khung hình",
                }
            ),
            200,
        )

    new_attendance = []
    dup_attendance = []
    emailed_attendance = []
    checked_at_now = datetime.now()
    lat_u, lng_u, loc_text = _coords_for_attendance(latitude, longitude)
    pending_new_codes: List[str] = []
    db_save_error: Optional[BaseException] = None
    conn = None
    try:
        conn = db_connect()
        conn.autocommit(False)
        with conn.cursor() as cur:
            for m in matches:
                sid, scode = _resolve_attendance_identity(m.get("student_db_id"), m.get("student_code"))
                # Không chặn theo student_subjects — cho điểm danh theo môn đã chọn (hoặc không môn).
                if _session_seen_today(scode):
                    dup_attendance.append(str(scode or ""))
                    continue
                if _attendance_exists_for_student_subject_today(cur, sid, scode, subject_id):
                    _mark_session_seen_today(scode)
                    dup_attendance.append(str(scode or ""))
                    continue
                cur.execute(
                    "INSERT INTO attendance (student_db_id, student_code, full_name, class_code, confidence, status, "
                    "latitude, longitude, location_text, subject_id) VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s,%s)",
                    (
                        sid,
                        scode,
                        m.get("full_name"),
                        m.get("class_code"),
                        float(m.get("confidence") or 0.0),
                        "success",
                        lat_u,
                        lng_u,
                        loc_text,
                        subject_id,
                    ),
                )
                if scode:
                    pending_new_codes.append(str(scode))
        conn.commit()
    except Exception as e:
        if conn is not None:
            try:
                conn.rollback()
            except Exception:
                pass
        db_save_error = e
        _log.exception("[ATTENDANCE] Lỗi khi lưu điểm danh sau nhận diện: %s", e)
    finally:
        if conn is not None:
            try:
                conn.autocommit(True)
            except Exception:
                pass
            try:
                conn.close()
            except Exception:
                pass

    if db_save_error is None:
        for scode in pending_new_codes:
            _mark_session_seen_today(scode)
            new_attendance.append(scode)
            try:
                _send_attendance_email_if_needed(
                    scode,
                    checked_at_now,
                    latitude=lat_u,
                    longitude=lng_u,
                )
                emailed_attendance.append(scode)
            except Exception as email_err:
                _log.error(f"[EMAIL] Send failed MSSV={scode}: {email_err}")

    if db_save_error is not None:
        return (
            jsonify(
                _json_fail(
                    "Lỗi khi lưu điểm danh",
                    {"success": False, "face_detected": True, "matches": matches, "count": len(matches)},
                )
            ),
            500,
        )

    for m in matches:
        _log_recognition(m.get("student_code"), float(m.get("confidence") or 0.0), True)

    first = matches[0]
    return (
        jsonify(
            {
                "ok": True,
                "success": True,
                "face_detected": True,
                "student_code": first.get("student_code"),
                "full_name": first.get("full_name"),
                "name": first.get("full_name"),
                "mssv": first.get("student_code"),
                "confidence": first.get("confidence"),
                "confidence_percent": first.get("confidence_percent"),
                "matches": matches,
                "count": len(matches),
                "new_attendance": new_attendance,
                "duplicate_attendance": dup_attendance,
                "emailed_attendance": emailed_attendance,
                "message": "OK",
            }
        ),
        200,
    )


def _haar_classifier() -> cv2.CascadeClassifier:
    global _haar
    if _haar is None:
        clf = cv2.CascadeClassifier(_CASCADE_PATH)
        if clf.empty():
            raise RuntimeError("Không load được Haar cascade. Kiểm tra cài đặt opencv-python.")
        _haar = clf
    return _haar


def _detect_and_crop_face_rgb(frame_rgb: np.ndarray) -> Optional[np.ndarray]:
    """
    Detect mặt lớn nhất bằng Haar, crop, trả về RGB crop (chưa resize) hoặc None.
    """
    if frame_rgb is None or frame_rgb.size == 0:
        return None
    frame_bgr = frame_rgb[:, :, ::-1].copy()
    gray = cv2.cvtColor(frame_bgr, cv2.COLOR_BGR2GRAY)
    gray = cv2.equalizeHist(gray)
    faces = _haar_classifier().detectMultiScale(gray, scaleFactor=1.1, minNeighbors=5, minSize=(48, 48))
    if faces is None or len(faces) == 0:
        return None
    x, y, w, h = max(faces, key=lambda r: int(r[2]) * int(r[3]))
    x, y, w, h = int(x), int(y), int(w), int(h)
    if w <= 0 or h <= 0:
        return None
    h_img, w_img = frame_bgr.shape[:2]
    x0 = max(0, x)
    y0 = max(0, y)
    x1 = min(w_img, x + w)
    y1 = min(h_img, y + h)
    crop_bgr = frame_bgr[y0:y1, x0:x1].copy()
    if crop_bgr.size == 0:
        return None
    return crop_bgr[:, :, ::-1].copy()


def _detect_crop_and_roi_rgb(frame_rgb: np.ndarray) -> Tuple[Optional[np.ndarray], Optional[Tuple[int, int, int, int]]]:
    """
    Detect mặt lớn nhất bằng Haar, trả về (crop_rgb, (x,y,w,h)) hoặc (None, None).
    """
    if frame_rgb is None or frame_rgb.size == 0:
        return None, None
    frame_bgr = frame_rgb[:, :, ::-1].copy()
    gray = cv2.cvtColor(frame_bgr, cv2.COLOR_BGR2GRAY)
    gray = cv2.equalizeHist(gray)
    faces = _haar_classifier().detectMultiScale(gray, scaleFactor=1.1, minNeighbors=5, minSize=(48, 48))
    if faces is None or len(faces) == 0:
        return None, None
    x, y, w, h = max(faces, key=lambda r: int(r[2]) * int(r[3]))
    x, y, w, h = int(x), int(y), int(w), int(h)
    if w <= 0 or h <= 0:
        return None, None
    h_img, w_img = frame_bgr.shape[:2]
    x0 = max(0, x)
    y0 = max(0, y)
    x1 = min(w_img, x + w)
    y1 = min(h_img, y + h)
    crop_bgr = frame_bgr[y0:y1, x0:x1].copy()
    if crop_bgr.size == 0:
        return None, (x, y, w, h)
    return crop_bgr[:, :, ::-1].copy(), (x, y, w, h)


def _embedding_model() -> tf.keras.Model:
    """
    Model tạo embedding. Ưu tiên FACENET_MODEL_PATH nếu có, fallback InceptionResNetV2 pooling='avg'.
    Không dùng tf.* trực tiếp trên KerasTensor (tránh lỗi KerasTensor).
    """
    global _embed_model
    if _embed_model is not None:
        return _embed_model
    model_path = os.environ.get("FACENET_MODEL_PATH") or ""
    if model_path and os.path.exists(model_path):
        _embed_model = tf.keras.models.load_model(model_path)
        return _embed_model
    _embed_model = tf.keras.applications.InceptionResNetV2(
        include_top=False,
        weights="imagenet",
        input_shape=(_FACE_SIZE[1], _FACE_SIZE[0], 3),
        pooling="avg",
    )
    return _embed_model


def _embedding_from_face_rgb(face_rgb: np.ndarray) -> np.ndarray:
    """
    Tạo embedding từ ảnh mặt RGB đã crop+resize.
    Dùng DeepFace để đồng bộ register/train/predict.
    """
    return embedding_from_face_rgb(face_rgb)


def _load_embeddings_db() -> Dict[str, Any]:
    """Gộp data/*/embeddings.npz (cùng logic nhận diện)."""
    return load_recognition_embeddings_db()


def _auth_serializer() -> URLSafeTimedSerializer:
    return URLSafeTimedSerializer(str(app.config["SECRET_KEY"]), salt="fa-admin-auth")


def create_auth_token(user_id: int, username: str) -> str:
    return _auth_serializer().dumps({"id": user_id, "u": username})


def verify_auth_token(token: str) -> Optional[Dict[str, Any]]:
    try:
        return _auth_serializer().loads(token, max_age=int(os.environ.get("AUTH_TOKEN_MAX_AGE", str(60 * 60 * 24 * 7))))
    except (BadSignature, SignatureExpired):
        return None


def _mysql_params() -> Dict[str, Any]:
    return {
        "host": os.environ.get("MYSQL_HOST", "localhost"),
        "port": int(os.environ.get("MYSQL_PORT", "3306")),
        "user": os.environ.get("MYSQL_USER", "root"),
        "password": os.environ.get("MYSQL_PASSWORD", ""),
        "database": os.environ.get("MYSQL_DATABASE", "face_attendance"),
        "charset": "utf8mb4",
        "cursorclass": pymysql.cursors.DictCursor,
        "autocommit": True,
    }


def db_connect():
    return pymysql.connect(**_mysql_params())


def init_schema() -> None:
    """Tạo bảng nếu chưa có (idempotent)."""
    ddl_classes = """
    CREATE TABLE IF NOT EXISTS classes (
      id INT AUTO_INCREMENT PRIMARY KEY,
      class_code VARCHAR(64) NOT NULL UNIQUE,
      class_name VARCHAR(255) NOT NULL,
      lecturer VARCHAR(255) NULL,
      notes TEXT NULL,
      created_at DATETIME DEFAULT CURRENT_TIMESTAMP
    ) ENGINE=InnoDB DEFAULT CHARSET=utf8mb4;
    """
    ddl_students = """
    CREATE TABLE IF NOT EXISTS students (
      id INT AUTO_INCREMENT PRIMARY KEY,
      student_code VARCHAR(64) NOT NULL UNIQUE,
      full_name VARCHAR(255) NOT NULL,
      email VARCHAR(255) NULL,
      phone VARCHAR(32) NULL,
      class_code VARCHAR(64) NULL,
      notes TEXT NULL,
      created_at DATETIME DEFAULT CURRENT_TIMESTAMP
    ) ENGINE=InnoDB DEFAULT CHARSET=utf8mb4;
    """
    ddl_admins = """
    CREATE TABLE IF NOT EXISTS admins (
      id INT AUTO_INCREMENT PRIMARY KEY,
      username VARCHAR(64) NOT NULL UNIQUE,
      password_hash VARCHAR(255) NOT NULL,
      created_at DATETIME DEFAULT CURRENT_TIMESTAMP
    ) ENGINE=InnoDB DEFAULT CHARSET=utf8mb4;
    """
    ddl_attendance = """
    CREATE TABLE IF NOT EXISTS attendance (
      id INT AUTO_INCREMENT PRIMARY KEY,
      student_db_id INT NULL,
      student_code VARCHAR(64) NULL,
      full_name VARCHAR(255) NULL,
      class_code VARCHAR(64) NULL,
      checked_at DATETIME DEFAULT CURRENT_TIMESTAMP,
      confidence DECIMAL(6,3) NULL,
      status VARCHAR(32) NOT NULL,
      latitude DECIMAL(10,6) NULL,
      longitude DECIMAL(10,6) NULL,
      location_text VARCHAR(512) NULL,
      CONSTRAINT fk_att_student FOREIGN KEY (student_db_id) REFERENCES students(id)
        ON DELETE SET NULL ON UPDATE CASCADE
    ) ENGINE=InnoDB DEFAULT CHARSET=utf8mb4;
    """
    ddl_subjects = """
    CREATE TABLE IF NOT EXISTS subjects (
      id INT AUTO_INCREMENT PRIMARY KEY,
      subject_code VARCHAR(64) NOT NULL UNIQUE,
      subject_name VARCHAR(255) NOT NULL,
      teacher VARCHAR(255) NULL,
      description TEXT NULL
    ) ENGINE=InnoDB DEFAULT CHARSET=utf8mb4;
    """
    ddl_student_subjects = """
    CREATE TABLE IF NOT EXISTS student_subjects (
      id INT AUTO_INCREMENT PRIMARY KEY,
      student_id INT NOT NULL,
      subject_id INT NOT NULL,
      UNIQUE KEY uq_student_subject (student_id, subject_id),
      CONSTRAINT fk_ss_student FOREIGN KEY (student_id) REFERENCES students(id)
        ON DELETE CASCADE ON UPDATE CASCADE,
      CONSTRAINT fk_ss_subject FOREIGN KEY (subject_id) REFERENCES subjects(id)
        ON DELETE CASCADE ON UPDATE CASCADE
    ) ENGINE=InnoDB DEFAULT CHARSET=utf8mb4;
    """
    ddl_settings = """
    CREATE TABLE IF NOT EXISTS settings (
      id INT PRIMARY KEY,
      enable_email BOOLEAN NOT NULL DEFAULT TRUE,
      gps VARCHAR(64) NULL
    ) ENGINE=InnoDB DEFAULT CHARSET=utf8mb4;
    """
    conn = db_connect()
    try:
        with conn.cursor() as cur:
            cur.execute(ddl_classes)
            cur.execute(ddl_students)
            cur.execute(ddl_admins)
            cur.execute(ddl_attendance)
            cur.execute(ddl_subjects)
            cur.execute(ddl_student_subjects)
            cur.execute(ddl_settings)
            # Bổ sung cột subject_id/class_id cho attendance ở các DB cũ.
            try:
                cur.execute("ALTER TABLE attendance ADD COLUMN subject_id INT NULL")
            except Exception:
                pass
            try:
                cur.execute("ALTER TABLE attendance ADD COLUMN class_id INT NULL")
            except Exception:
                pass
            # Index tối ưu truy vấn xem điểm danh theo môn/ngày.
            try:
                cur.execute("CREATE INDEX idx_att_subject_date_student ON attendance(subject_id, checked_at, student_db_id)")
            except Exception:
                pass
            try:
                cur.execute("CREATE INDEX idx_ss_subject_student ON student_subjects(subject_id, student_id)")
            except Exception:
                pass
            cur.execute("SELECT id FROM settings WHERE id=1 LIMIT 1")
            if not cur.fetchone():
                cur.execute(
                    "INSERT INTO settings (id, enable_email, gps) VALUES (1, %s, %s)",
                    (1, None),
                )
            cur.execute("SELECT COUNT(*) AS c FROM admins")
            if int(cur.fetchone()["c"]) == 0:
                default_pw = os.environ.get("ADMIN_DEFAULT_PASSWORD", "123")
                cur.execute(
                    "INSERT INTO admins (username, password_hash) VALUES (%s,%s)",
                    ("admin", generate_password_hash(default_pw)),
                )
    finally:
        conn.close()


def fetch_student_by_code(student_code: str) -> Optional[Dict[str, Any]]:
    conn = db_connect()
    try:
        with conn.cursor() as cur:
            cur.execute(
                "SELECT id, student_code, full_name, email, phone, class_code, notes, created_at "
                "FROM students WHERE student_code=%s LIMIT 1",
                (student_code,),
            )
            return cur.fetchone()
    finally:
        conn.close()


def get_settings() -> Dict[str, Any]:
    """
    Đọc setting hệ thống (record duy nhất id=1).
    Nếu chưa có thì tự tạo default.
    """
    conn = db_connect()
    try:
        with conn.cursor() as cur:
            cur.execute("SELECT id, enable_email, gps FROM settings WHERE id=1 LIMIT 1")
            row = cur.fetchone()
            if not row:
                cur.execute(
                    "INSERT INTO settings (id, enable_email, gps) VALUES (1, %s, %s)",
                    (1, None),
                )
                cur.execute("SELECT id, enable_email, gps FROM settings WHERE id=1 LIMIT 1")
                row = cur.fetchone()
            return {
                "id": 1,
                "enable_email": bool(row.get("enable_email")) if row else True,
                "gps": (str((row or {}).get("gps") or "").strip() or None),
            }
    finally:
        conn.close()


def _json_fail(message: str, extra: Optional[Dict[str, Any]] = None) -> Dict[str, Any]:
    """Payload JSON lỗi chuẩn (status/message) — không xóa field cũ nếu truyền thêm qua extra."""
    d: Dict[str, Any] = {"ok": False, "status": "fail", "message": message}
    if extra:
        d.update(extra)
    return d


def _parse_saved_gps(gps_text: Optional[str]) -> Tuple[Optional[float], Optional[float]]:
    if not gps_text:
        return None, None
    raw = str(gps_text).strip()
    if not raw:
        return None, None
    parts = [p.strip() for p in raw.split(",")]
    if len(parts) != 2:
        return None, None
    try:
        lat = float(parts[0])
        lon = float(parts[1])
        return lat, lon
    except Exception:
        return None, None


def _parse_subject_ids(raw: Any) -> List[int]:
    """Danh sách môn từ JSON (đăng ký/cập nhật sinh viên)."""
    if raw is None:
        return []
    vals = raw if isinstance(raw, list) else [raw]
    out: List[int] = []
    for v in vals:
        try:
            iv = int(v)
            if iv > 0 and iv not in out:
                out.append(iv)
        except Exception:
            continue
    return out


def _set_student_subjects(cur, student_id: int, subject_ids: List[int]) -> None:
    cur.execute("DELETE FROM student_subjects WHERE student_id=%s", (student_id,))
    for sid in subject_ids:
        cur.execute(
            "INSERT INTO student_subjects (student_id, subject_id) VALUES (%s,%s)",
            (student_id, sid),
        )


def _parse_request_gps_coordinate(value: Any) -> Optional[float]:
    """
    Đọc latitude/longitude từ JSON (số hoặc chuỗi). None/"" → None.
    0 và 0.0 là tọa độ hợp lệ. Loại bool, NaN, Inf.
    """
    if value is None:
        return None
    if isinstance(value, bool):
        return None
    if isinstance(value, str):
        s = value.strip()
        if not s:
            return None
        try:
            v = float(s)
        except ValueError:
            return None
    else:
        try:
            v = float(value)
        except (TypeError, ValueError):
            return None
    if math.isnan(v) or math.isinf(v):
        return None
    return float(v)


def _coords_for_attendance(latitude: Optional[float], longitude: Optional[float]) -> Tuple[float, float, str]:
    """
    Luôn trả về tọa độ + chuỗi hiển thị GPS để lưu attendance (không để trống location).
    Ưu tiên tọa độ client; fallback GPS trong settings; cuối cùng 0,0 + nhãn.
    """
    if latitude is not None and longitude is not None:
        return (
            float(latitude),
            float(longitude),
            f"{float(latitude):.6f}, {float(longitude):.6f}",
        )
    cfg = get_settings()
    slat, slng = _parse_saved_gps(cfg.get("gps"))
    if slat is not None and slng is not None:
        return (
            float(slat),
            float(slng),
            f"{float(slat):.6f}, {float(slng):.6f}",
        )
    return 0.0, 0.0, "0.000000, 0.000000 (chưa định vị)"


def _location_text_same_as_coords(lat: Any, lng: Any, location_text: str) -> bool:
    """
    Khi lưu attendance, location_text thường trùng nội dung với latitude/longitude.
    Tránh ghép 2 lần trên API lịch sử.
    """
    if not location_text or lat is None or lng is None:
        return False
    nums = re.findall(r"-?\d+\.?\d*", str(location_text))
    if len(nums) != 2:
        return False
    try:
        return abs(float(nums[0]) - float(lat)) < 1e-4 and abs(float(nums[1]) - float(lng)) < 1e-4
    except Exception:
        return False


def _format_attendance_location_display(lat: Any, lng: Any, location_text: Optional[str]) -> str:
    """Một dòng hiển thị: tọa độ; hoặc tọa độ · ghi chú nếu location_text khác tọa độ.
    Luôn coi tọa độ 0,0 là hợp lệ (không dùng truthy `if lat and lng` — 0 bị bỏ sót).
    """
    coord_str: Optional[str] = None
    if lat is not None and lng is not None:
        try:
            coord_str = f"{float(lat):.6f}, {float(lng):.6f}"
        except (TypeError, ValueError):
            coord_str = f"{lat}, {lng}"
    lt = str(location_text or "").strip()
    if not coord_str and not lt:
        return "—"
    if coord_str and lt:
        if _location_text_same_as_coords(lat, lng, lt):
            return coord_str
        if "".join(coord_str.split()) == "".join(lt.split()):
            return coord_str
        return f"{coord_str} · {lt}"
    return coord_str or lt or "—"


def _haversine_distance_meters(lat1: float, lon1: float, lat2: float, lon2: float) -> float:
    """
    Tính khoảng cách giữa 2 tọa độ GPS theo công thức Haversine.
    """
    r = 6371000.0  # bán kính trái đất (m)
    p1 = math.radians(lat1)
    p2 = math.radians(lat2)
    dlat = math.radians(lat2 - lat1)
    dlon = math.radians(lon2 - lon1)
    a = math.sin(dlat / 2) ** 2 + math.cos(p1) * math.cos(p2) * math.sin(dlon / 2) ** 2
    c = 2 * math.atan2(math.sqrt(a), math.sqrt(1 - a))
    return r * c


def _send_attendance_email_if_needed(
    student_code: Optional[str],
    checked_at: datetime,
    latitude: Optional[float] = None,
    longitude: Optional[float] = None,
) -> None:
    """
    Gửi email thông báo khi điểm danh thành công.
    Hàm này được gọi sau khi insert attendance mới để tránh gửi trùng.
    """
    scode = (str(student_code).strip() if student_code is not None else "")
    if not scode:
        return
    cfg = get_settings()
    if not bool(cfg.get("enable_email", True)):
        _log.info(f"[EMAIL] Disabled by settings. Skip MSSV={scode}.")
        return

    student = fetch_student_by_code(scode)
    if not student:
        _log.warning(f"[EMAIL] Student not found for MSSV={scode}.")
        return

    to_email = str(student.get("email") or "").strip()
    if not to_email:
        _log.warning(f"[EMAIL] Missing email for MSSV={scode}.")
        return

    name = str(student.get("full_name") or scode)
    checked_at_str = checked_at.strftime("%Y-%m-%d %H:%M:%S")
    gps_text = "Chưa cập nhật"
    if latitude is not None and longitude is not None:
        gps_text = f"{latitude:.6f}, {longitude:.6f}"
    send_email_async(
        to_email=to_email,
        mssv=scode,
        name=name,
        time=checked_at_str,
        status="Có mặt",
        gps=gps_text,
    )
    _log.info(f"[EMAIL] Queued attendance email MSSV={scode} -> {to_email}")


@app.route("/health", methods=["GET"])
def health():
    # DB health
    db_status = "ok"
    try:
        conn = db_connect()
        try:
            with conn.cursor() as cur:
                cur.execute("SELECT 1 AS ok")
                _ = cur.fetchone()
        finally:
            conn.close()
    except Exception:
        db_status = "error"

    # Facenet health (giữ nhẹ, không đổi luồng model hiện tại)
    facenet_status = "ok"

    # YOLO health
    yolo_info = get_yolo_status()
    if yolo_info.get("status") != "ok" and not yolo_info.get("attempted"):
        try:
            get_yolo_detector()
        except Exception:
            pass
        yolo_info = get_yolo_status()
    yolo_status = str(yolo_info.get("status") or "error")

    return jsonify(
        {
            "status": "ok",
            "service": "face-attendance-api",
            "yolo": yolo_status,
            "facenet": facenet_status,
            "database": db_status,
        }
    )


@app.route("/settings", methods=["GET"])
def settings_get():
    cfg = get_settings()
    return jsonify({"ok": True, "settings": cfg}), 200


@app.route("/settings", methods=["POST"])
def settings_save():
    body = request.get_json(silent=True) or {}
    enable_email = body.get("enable_email")
    gps = body.get("gps")
    # Parse boolean rõ ràng để tránh trường hợp chuỗi "false" bị hiểu thành True.
    if isinstance(enable_email, bool):
        enable_email_bool = enable_email
    elif isinstance(enable_email, (int, float)):
        enable_email_bool = bool(int(enable_email))
    else:
        enable_email_bool = str(enable_email).strip().lower() in ("1", "true", "yes", "on")
    gps_text = str(gps).strip() if gps is not None else None
    if gps_text == "":
        gps_text = None
    conn = db_connect()
    try:
        with conn.cursor() as cur:
            cur.execute(
                "INSERT INTO settings (id, enable_email, gps) VALUES (1, %s, %s) "
                "ON DUPLICATE KEY UPDATE enable_email=VALUES(enable_email), gps=VALUES(gps)",
                (enable_email_bool, gps_text),
            )
        return jsonify({"ok": True, "settings": {"id": 1, "enable_email": enable_email_bool, "gps": gps_text}}), 200
    finally:
        conn.close()


def _remove_legacy_root_embeddings_npz(root: Path) -> None:
    """File embeddings.npz cũ (một file chung) nếu còn sót."""
    p = root / "embeddings.npz"
    if p.is_file():
        try:
            p.unlink()
        except Exception:
            pass


def _clear_entire_data_directory(root: Path) -> None:
    """Xóa toàn bộ nội dung trong thư mục data/."""
    if not root.is_dir():
        return
    for child in list(root.iterdir()):
        try:
            if child.is_dir():
                shutil.rmtree(child, ignore_errors=True)
            else:
                child.unlink(missing_ok=True)
        except Exception:
            pass


def _delete_all_student_folders(root: Path) -> None:
    """Xóa từng thư mục con data/<MSSV>/ và file embeddings gốc cũ."""
    if not root.is_dir():
        return
    for child in list(root.iterdir()):
        if child.is_dir() and not child.name.startswith("."):
            try:
                shutil.rmtree(child, ignore_errors=True)
            except Exception:
                pass
    _remove_legacy_root_embeddings_npz(root)


@app.route("/settings/wipe", methods=["POST"])
def settings_wipe():
    """
    Xóa dữ liệu theo phạm vi (cần xác nhận ở client).
    Body JSON: { "scope": "classes" | "students" | "full" }
    - classes: DELETE FROM classes
    - students: xóa attendance + student_subjects + students + thư mục data/<MSSV>/
    - full: xóa toàn bộ bảng liên quan + toàn bộ nội dung data/
    """
    body = request.get_json(silent=True) or {}
    scope = (body.get("scope") or "").strip().lower()
    root = data_dir()
    try:
        conn = db_connect()
        try:
            with conn.cursor() as cur:
                if scope == "classes":
                    cur.execute("DELETE FROM classes")
                elif scope == "students":
                    cur.execute("DELETE FROM attendance")
                    cur.execute("DELETE FROM student_subjects")
                    cur.execute("DELETE FROM students")
                    _delete_all_student_folders(root)
                elif scope == "full":
                    cur.execute("DELETE FROM attendance")
                    cur.execute("DELETE FROM student_subjects")
                    cur.execute("DELETE FROM students")
                    cur.execute("DELETE FROM classes")
                    cur.execute("DELETE FROM subjects")
                    _clear_entire_data_directory(root)
                else:
                    return (
                        jsonify(
                            {
                                "ok": False,
                                "success": False,
                                "error": "scope không hợp lệ. Dùng: classes, students, full.",
                            }
                        ),
                        400,
                    )
        finally:
            conn.close()
        reset_embeddings_cache()
        return jsonify({"ok": True, "success": True, "scope": scope}), 200
    except Exception as e:
        _log.error(f"[WIPE] scope={scope} error: {e}")
        return jsonify({"ok": False, "success": False, "error": str(e)}), 500


@app.route("/email/test", methods=["POST"])
def email_test():
    """
    Test SMTP gửi email.
    Body JSON (optional):
    - to_email: email nhận test trực tiếp
    - student_code: lấy email theo MSSV trong bảng students
    """
    body = request.get_json(silent=True) or {}
    to_email = str(body.get("to_email") or "").strip()
    student_code = str(body.get("student_code") or "").strip()

    if not to_email and student_code:
        stu = fetch_student_by_code(student_code)
        if stu:
            to_email = str(stu.get("email") or "").strip()

    if not to_email:
        return (
            jsonify(
                {
                    "ok": False,
                    "error": "Thiếu to_email hoặc student_code hợp lệ có email.",
                }
            ),
            400,
        )

    now_str = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    mssv = student_code or "TEST001"
    name = str(body.get("name") or "Sinh viên test")
    status = str(body.get("status") or "Có mặt")

    try:
        sent = send_email(
            to_email=to_email,
            mssv=mssv,
            name=name,
            time=now_str,
            status=status,
        )
        if not sent:
            return jsonify({"ok": False, "error": "SMTP gửi thất bại. Kiểm tra cấu hình .env."}), 500
        return jsonify({"ok": True, "message": "Gửi email test thành công.", "to_email": to_email}), 200
    except Exception as e:
        return jsonify({"ok": False, "error": f"Gửi email test lỗi: {e}"}), 500


@app.route("/dashboard", methods=["GET"])
def dashboard():
    """
    Dashboard stats lấy từ MySQL (dữ liệu thật).
    """
    conn = db_connect()
    try:
        with conn.cursor() as cur:
            cur.execute("SELECT COUNT(*) AS c FROM students")
            total_students = int(cur.fetchone()["c"])
            cur.execute("SELECT COUNT(*) AS c FROM classes")
            total_classes = int(cur.fetchone()["c"])
            cur.execute(
                "SELECT COUNT(DISTINCT a.student_db_id) AS c "
                "FROM attendance a "
                "WHERE DATE(checked_at) = CURDATE() "
                "AND a.student_db_id IS NOT NULL "
                "AND LOWER(COALESCE(a.status,'')) = 'success'"
            )
            total_attendance_today = int(cur.fetchone()["c"])
            # Chống % nhảy sai do duplicate record trong cùng 1 ngày:
            # - total_days: số ngày có dữ liệu điểm danh (distinct date)
            # - present_days: số ngày có trạng thái present/success (distinct date)
            cur.execute(
                "SELECT "
                "COUNT(DISTINCT DATE(checked_at)) AS total_days, "
                "COUNT(DISTINCT CASE "
                "  WHEN LOWER(COALESCE(status, '')) IN ('present', 'success') "
                "  THEN DATE(checked_at) "
                "END) AS present_days "
                "FROM attendance"
            )
            day_stats = cur.fetchone() or {}
            total_days = int(day_stats.get("total_days") or 0)
            present_days = int(day_stats.get("present_days") or 0)
            cur.execute(
                "SELECT COALESCE(s.full_name, a.full_name) AS full_name, "
                "a.checked_at AS checked_at, "
                "COALESCE(c.class_name, s.class_code, a.class_code) AS class_name, "
                "a.status AS status "
                "FROM attendance a "
                "JOIN ("
                "  SELECT "
                "    COALESCE(a0.student_code, CONCAT('db:', COALESCE(a0.student_db_id, 0))) AS ident, "
                "    DATE(a0.checked_at) AS day_key, "
                "    MAX(a0.id) AS max_id "
                "  FROM attendance a0 "
                "  JOIN students s0 ON (a0.student_db_id = s0.id OR a0.student_code = s0.student_code) "
                "  GROUP BY ident, day_key "
                "  ORDER BY max_id DESC "
                "  LIMIT 10"
                ") t ON a.id = t.max_id "
                "JOIN students s ON (a.student_db_id = s.id OR a.student_code = s.student_code) "
                "LEFT JOIN classes c ON (s.class_code = c.class_code) "
                "ORDER BY a.checked_at DESC, a.id DESC "
                "LIMIT 10"
            )
            rows = cur.fetchall()
    finally:
        conn.close()

    recent = []
    for r in rows:
        t = r.get("checked_at")
        if isinstance(t, datetime):
            t_str = t.strftime("%Y-%m-%d %H:%M:%S")
        else:
            t_str = str(t) if t is not None else None
        recent.append(
            {
                "name": r.get("full_name"),
                "time": t_str,
                "class": r.get("class_name"),
                "status": r.get("status"),
            }
        )

    percentage = round((present_days / total_days) * 100.0, 2) if total_days > 0 else 0.0
    return (
        jsonify(
            {
                "success": True,
                "data": {
                    "total_students": total_students,
                    "total_classes": total_classes,
                    "total_attendance_today": total_attendance_today,
                    "attendance_rate": percentage,
                    "total_days": total_days,
                    "present_days": present_days,
                    "percentage": percentage,
                    "recent_attendance": recent,
                },
            }
        ),
        200,
    )


@app.route("/auth/login", methods=["GET", "POST"])
def auth_login():
    if request.method == "GET":
        return jsonify(
            {
                "ok": False,
                "message": "Đăng nhập API chỉ dùng POST với JSON {username, password}. "
                "Mở trang đăng nhập qua giao diện web (ví dụ http://localhost:8080/mnm/), không gõ URL /auth/login trên thanh địa chỉ.",
            }
        ), 200
    body = request.get_json(silent=True) or {}
    username = (body.get("username") or "").strip()
    password = body.get("password") or ""
    if not username:
        _log_login(False, username, reason="missing_username")
        return jsonify({"ok": False, "error": "Thiếu tên đăng nhập."}), 400
    conn = db_connect()
    try:
        with conn.cursor() as cur:
            cur.execute("SELECT id, username, password_hash FROM admins WHERE username=%s LIMIT 1", (username,))
            row = cur.fetchone()
        if not row or not check_password_hash(row["password_hash"], password):
            _log_login(False, username, reason="invalid_credentials")
            return jsonify({"ok": False, "error": "Sai tài khoản hoặc mật khẩu."}), 401
        token = create_auth_token(int(row["id"]), row["username"])
        _log_login(True, username)
        return jsonify({"ok": True, "token": token, "username": row["username"]}), 200
    finally:
        conn.close()


@app.route("/auth/me", methods=["GET"])
def auth_me():
    adm = getattr(g, "admin", None) or {}
    return jsonify({"ok": True, "username": adm.get("u")}), 200


@app.route("/students/check", methods=["GET"])
def students_check():
    code = (request.args.get("student_code") or "").strip()
    if not code:
        return jsonify({"ok": False, "error": "Thiếu student_code."}), 400
    data_folder = data_dir() / code
    data_folder_exists = data_folder.is_dir()
    conn = db_connect()
    try:
        with conn.cursor() as cur:
            cur.execute("SELECT 1 FROM students WHERE student_code=%s LIMIT 1", (code,))
            exists = cur.fetchone() is not None
        return jsonify({"ok": True, "exists": exists, "data_folder_exists": data_folder_exists}), 200
    finally:
        conn.close()


@app.route("/classes/check", methods=["GET"])
def classes_check():
    code = (request.args.get("class_code") or "").strip()
    if not code:
        return jsonify({"ok": False, "error": "Thiếu class_code."}), 400
    conn = db_connect()
    try:
        with conn.cursor() as cur:
            cur.execute("SELECT 1 FROM classes WHERE class_code=%s LIMIT 1", (code,))
            exists = cur.fetchone() is not None
        return jsonify({"ok": True, "exists": exists}), 200
    finally:
        conn.close()


@app.route("/classes", methods=["GET"])
def classes_list():
    """Danh sách lớp kèm sĩ số (đếm sinh viên theo class_code)."""
    conn = db_connect()
    try:
        with conn.cursor() as cur:
            cur.execute(
                "SELECT c.id, c.class_code, c.class_name, c.lecturer, c.notes, c.created_at, "
                "(SELECT COUNT(*) FROM students s WHERE s.class_code = c.class_code) AS student_count "
                "FROM classes c ORDER BY c.class_code ASC"
            )
            rows = cur.fetchall()
        for r in rows:
            if isinstance(r.get("created_at"), datetime):
                r["created_at"] = r["created_at"].strftime("%Y-%m-%d %H:%M:%S")
            r["student_count"] = int(r.get("student_count") or 0)
        return jsonify({"ok": True, "classes": rows}), 200
    finally:
        conn.close()


@app.route("/classes", methods=["POST"])
def classes_create():
    body = request.get_json(silent=True) or {}
    code = (body.get("class_code") or "").strip()
    name = (body.get("class_name") or "").strip()
    if not code or not name:
        return jsonify({"ok": False, "error": "class_code và class_name là bắt buộc."}), 400
    lecturer = body.get("lecturer")
    notes = body.get("notes")
    conn = db_connect()
    try:
        with conn.cursor() as cur:
            cur.execute(
                "INSERT INTO classes (class_code, class_name, lecturer, notes) VALUES (%s,%s,%s,%s)",
                (code, name, lecturer, notes),
            )
            new_id = cur.lastrowid
        return jsonify({"ok": True, "id": new_id}), 201
    except pymysql.err.IntegrityError:
        return jsonify({"ok": False, "error": "Mã lớp đã tồn tại."}), 409
    finally:
        conn.close()


@app.route("/classes/<int:cid>", methods=["PUT"])
def classes_update(cid: int):
    body = request.get_json(silent=True) or {}
    fields = []
    params: List[Any] = []
    for k, col in (("class_name", "class_name"), ("lecturer", "lecturer"), ("notes", "notes")):
        if k in body:
            fields.append(f"{col}=%s")
            params.append(body[k])
    if not fields:
        return jsonify({"ok": False, "error": "Không có trường cập nhật."}), 400
    params.append(cid)
    sql = "UPDATE classes SET " + ", ".join(fields) + " WHERE id=%s"
    conn = db_connect()
    try:
        with conn.cursor() as cur:
            cur.execute(sql, tuple(params))
            if cur.rowcount == 0:
                return jsonify({"ok": False, "error": "Không tìm thấy lớp."}), 404
        return jsonify({"ok": True}), 200
    finally:
        conn.close()


def _delete_student_row_and_data(student_code: str) -> None:
    import shutil

    folder = data_dir() / student_code
    if folder.is_dir():
        shutil.rmtree(folder, ignore_errors=True)


@app.route("/classes/<int:cid>", methods=["DELETE"])
def classes_delete(cid: int):
    """
    Xóa lớp. Nếu còn sinh viên: trả 409 trừ khi ?delete_students=1 (xoá hết SV thuộc lớp + ảnh train).
    """
    delete_students = request.args.get("delete_students") == "1"
    conn = db_connect()
    try:
        with conn.cursor() as cur:
            cur.execute("SELECT id, class_code FROM classes WHERE id=%s", (cid,))
            row = cur.fetchone()
            if not row:
                return jsonify({"ok": False, "error": "Không tìm thấy lớp."}), 404
            class_code = row["class_code"]
            cur.execute("SELECT COUNT(*) AS c FROM students WHERE class_code=%s", (class_code,))
            cnt = int(cur.fetchone()["c"])
            if cnt > 0 and not delete_students:
                return (
                    jsonify(
                        {
                            "ok": False,
                            "error": "Lớp còn sinh viên. Thêm ?delete_students=1 hoặc xóa SV trước.",
                            "student_count": cnt,
                        }
                    ),
                    409,
                )
            if cnt > 0 and delete_students:
                cur.execute(
                    "SELECT id, student_code FROM students WHERE class_code=%s",
                    (class_code,),
                )
                studs = cur.fetchall()
                for s in studs:
                    sid = int(s["id"])
                    scode = s["student_code"]
                    cur.execute("DELETE FROM students WHERE id=%s", (sid,))
                    _delete_student_row_and_data(scode)
            cur.execute("DELETE FROM classes WHERE id=%s", (cid,))
        return jsonify({"ok": True, "deleted_students": cnt if delete_students else 0}), 200
    finally:
        conn.close()


@app.route("/subjects", methods=["GET"])
def subjects_list():
    conn = db_connect()
    try:
        with conn.cursor() as cur:
            cur.execute(
                "SELECT s.id, s.subject_code, s.subject_name, s.teacher, s.description, "
                "(SELECT COUNT(*) FROM student_subjects ss WHERE ss.subject_id=s.id) AS student_count "
                "FROM subjects s ORDER BY s.subject_code ASC"
            )
            rows = cur.fetchall() or []
        for r in rows:
            r["student_count"] = int(r.get("student_count") or 0)
        return jsonify({"ok": True, "subjects": rows}), 200
    finally:
        conn.close()


@app.route("/subjects", methods=["POST"])
def subjects_create():
    body = request.get_json(silent=True) or {}
    code = str(body.get("subject_code") or "").strip()
    name = str(body.get("subject_name") or "").strip()
    teacher = body.get("teacher")
    description = body.get("description")
    if not code or not name:
        return jsonify({"ok": False, "error": "subject_code và subject_name là bắt buộc."}), 400
    conn = db_connect()
    try:
        with conn.cursor() as cur:
            cur.execute(
                "INSERT INTO subjects (subject_code, subject_name, teacher, description) VALUES (%s,%s,%s,%s)",
                (code, name, teacher, description),
            )
            new_id = int(cur.lastrowid)
        return jsonify({"ok": True, "id": new_id}), 201
    except pymysql.err.IntegrityError:
        return jsonify({"ok": False, "error": "Mã môn học đã tồn tại."}), 409
    finally:
        conn.close()


@app.route("/subjects/<int:sid>", methods=["PUT"])
def subjects_update(sid: int):
    body = request.get_json(silent=True) or {}
    fields = []
    params: List[Any] = []
    for k, col in (
        ("subject_code", "subject_code"),
        ("subject_name", "subject_name"),
        ("teacher", "teacher"),
        ("description", "description"),
    ):
        if k in body:
            fields.append(f"{col}=%s")
            params.append(body[k])
    if not fields:
        return jsonify({"ok": False, "error": "Không có trường cập nhật."}), 400
    params.append(sid)
    conn = db_connect()
    try:
        with conn.cursor() as cur:
            cur.execute("UPDATE subjects SET " + ", ".join(fields) + " WHERE id=%s", tuple(params))
            if cur.rowcount == 0:
                return jsonify({"ok": False, "error": "Không tìm thấy môn học."}), 404
        return jsonify({"ok": True}), 200
    except pymysql.err.IntegrityError:
        return jsonify({"ok": False, "error": "Mã môn học đã tồn tại."}), 409
    finally:
        conn.close()


@app.route("/subjects/<int:sid>", methods=["DELETE"])
def subjects_delete(sid: int):
    conn = db_connect()
    try:
        with conn.cursor() as cur:
            cur.execute("DELETE FROM subjects WHERE id=%s", (sid,))
            if cur.rowcount == 0:
                return jsonify({"ok": False, "error": "Không tìm thấy môn học."}), 404
        return jsonify({"ok": True}), 200
    finally:
        conn.close()


@app.route("/subjects/<int:subject_id>/attendance", methods=["GET"])
def subject_attendance_list(subject_id: int):
    """
    Danh sách SV thuộc môn + trạng thái điểm danh theo ngày.
    Query: date=YYYY-MM-DD (optional, mặc định hôm nay)
    """
    date_str = (request.args.get("date") or "").strip()
    if not date_str:
        date_str = datetime.now().strftime("%Y-%m-%d")
    conn = db_connect()
    try:
        with conn.cursor() as cur:
            cur.execute(
                "SELECT id, subject_code, subject_name, teacher FROM subjects WHERE id=%s LIMIT 1",
                (subject_id,),
            )
            subject = cur.fetchone()
            if not subject:
                return jsonify({"ok": False, "error": "Không tìm thấy môn học."}), 404

            # LEFT JOIN attendance: bản ghi mới nhất trong ngày theo môn + khớp student_db_id hoặc student_code
            cur.execute(
                "SELECT s.id AS student_id, s.student_code, s.full_name, s.class_code, "
                "c.class_name, a.checked_at, a.status "
                "FROM student_subjects ss "
                "JOIN students s ON s.id = ss.student_id "
                "LEFT JOIN classes c ON c.class_code = s.class_code "
                "LEFT JOIN attendance a ON a.id = ("
                "  SELECT MAX(a2.id) FROM attendance a2 "
                "  WHERE a2.subject_id=%s AND DATE(a2.checked_at)=%s "
                "  AND ("
                "    (a2.student_db_id IS NOT NULL AND a2.student_db_id = s.id) "
                "    OR (a2.student_code IS NOT NULL AND a2.student_code = s.student_code)"
                "  )"
                ") "
                "WHERE ss.subject_id=%s "
                "ORDER BY s.student_code ASC",
                (subject_id, date_str, subject_id),
            )
            rows = cur.fetchall() or []
        items: List[Dict[str, Any]] = []
        for r in rows:
            checked = r.get("checked_at")
            checked_text = checked.strftime("%Y-%m-%d %H:%M:%S") if isinstance(checked, datetime) else None
            st = str(r.get("status") or "").lower()
            is_present = checked is not None and st in ("success", "present")
            items.append(
                {
                    "student_id": int(r.get("student_id") or 0),
                    "mssv": r.get("student_code"),
                    "name": r.get("full_name"),
                    "class": r.get("class_name") or r.get("class_code"),
                    "status": "Đã điểm danh" if is_present else "Chưa điểm danh",
                    "time": checked_text,
                    "present": bool(is_present),
                }
            )
        return jsonify({"ok": True, "subject": subject, "date": date_str, "items": items}), 200
    finally:
        conn.close()


@app.route("/train", methods=["POST"])
def train():
    """
    Train CNN từ thư mục data/<MSSV>/*.jpg
    Body JSON:
    - { "epochs": 30, "mode": "all" }
    - { "epochs": 30, "mode": "class", "class_id": 1 }
    """
    body = request.get_json(silent=True) or {}
    epochs = int(body.get("epochs", default_train_epochs()))
    mode = str(body.get("mode") or "all").strip().lower() or "all"
    class_id = body.get("class_id")
    try:
        allowed_codes: Optional[Set[str]] = None
        if mode == "class":
            if class_id is None or str(class_id).strip() == "":
                return jsonify({"ok": False, "error": "Thiếu class_id khi mode=class."}), 400
            conn = db_connect()
            try:
                with conn.cursor() as cur:
                    cur.execute("SELECT class_code FROM classes WHERE id=%s LIMIT 1", (class_id,))
                    cls = cur.fetchone()
                    if not cls:
                        return jsonify({"ok": False, "error": "Không tìm thấy lớp theo class_id."}), 404
                    class_code = str(cls.get("class_code") or "").strip()
                    cur.execute("SELECT student_code FROM students WHERE class_code=%s", (class_code,))
                    rows = cur.fetchall() or []
                    allowed_codes = {
                        str(r.get("student_code") or "").strip() for r in rows if str(r.get("student_code") or "").strip()
                    }
            finally:
                conn.close()
            if not allowed_codes:
                return jsonify({"ok": False, "error": "Lớp không có sinh viên để train."}), 400
        elif mode != "all":
            return jsonify({"ok": False, "error": "mode không hợp lệ. Dùng 'all' hoặc 'class'."}), 400

        result = train_from_directory(epochs=epochs, allowed_student_codes=allowed_codes)
        # Chuẩn hóa key cho frontend: luôn có total_images.
        if "total_images" not in result:
            result["total_images"] = int(result.get("num_images") or result.get("images_count") or 0)
        result["mode"] = mode
        if mode == "class":
            result["class_id"] = int(class_id)
        # đảm bảo train lại toàn bộ và dùng embeddings mới ngay
        reset_embeddings_cache()
        return jsonify(result), 200
    except ValueError as e:
        return jsonify({"ok": False, "error": str(e)}), 400
    except Exception as e:
        return jsonify({"ok": False, "error": str(e)}), 500


@app.route("/predict", methods=["POST"])
def predict():
    """
    Nhận 1 ảnh (multipart file `image` hoặc base64), trả MSSV + confidence.
    Body JSON: { "image": "<base64 hoặc data URL>" } hoặc { "images": ["<base64>", ...] } (lấy ảnh đầu)
    """
    body = request.get_json(silent=True) or {}
    # tuỳ chọn: nhận GPS từ client để lưu vào lịch sử
    req_lat = _parse_request_gps_coordinate(body.get("latitude"))
    req_lng = _parse_request_gps_coordinate(body.get("longitude"))
    req_subject_id = body.get("subject_id")
    try:
        req_subject_id = int(req_subject_id) if req_subject_id is not None and str(req_subject_id).strip() != "" else None
    except Exception:
        req_subject_id = None

    # Check vị trí điểm danh theo GPS đã cấu hình (nếu có)
    cfg = get_settings()
    saved_lat, saved_lng = _parse_saved_gps(cfg.get("gps"))
    if saved_lat is not None and saved_lng is not None and req_lat is not None and req_lng is not None:
        dist_m = _haversine_distance_meters(req_lat, req_lng, saved_lat, saved_lng)
        if dist_m > 50.0:
            return (
                jsonify(
                    {
                        "ok": True,
                        "status": "fail",
                        "success": False,
                        "message": "Vui lòng ở đúng vị trí để điểm danh",
                        "distance_meters": round(dist_m, 2),
                    }
                ),
                200,
            )

    frame_rgb = _read_frame_rgb_from_request_json_or_multipart(body)
    if frame_rgb is None:
        return jsonify({"ok": False, "error": "Thiếu ảnh: upload file `image` hoặc JSON `image`/`images`."}), 400
    return _recognize_and_attend_from_frame_rgb(
        frame_rgb=frame_rgb,
        latitude=req_lat,
        longitude=req_lng,
        subject_id=req_subject_id,
    )


@app.route("/students", methods=["GET"])
def students_list():
    class_code = request.args.get("class_code")
    class_id = (request.args.get("class_id") or "").strip()
    conn = db_connect()
    base_select = (
        "SELECT s.id, s.student_code, s.full_name, s.email, s.phone, s.class_code, s.notes, s.created_at, "
        "EXISTS ("
        " SELECT 1 FROM attendance a WHERE "
        " (a.student_code = s.student_code OR a.student_db_id = s.id) "
        " AND DATE(a.checked_at) = CURDATE() "
        " AND LOWER(COALESCE(a.status,'')) = 'success'"
        ") AS attended_today "
        "FROM students s "
    )
    try:
        with conn.cursor() as cur:
            if class_id and class_id.lower() != "all":
                cur.execute("SELECT class_code FROM classes WHERE id=%s LIMIT 1", (class_id,))
                cls = cur.fetchone()
                selected_class_code = str((cls or {}).get("class_code") or "").strip()
                if selected_class_code:
                    cur.execute(
                        base_select + "WHERE s.class_code=%s ORDER BY s.id DESC",
                        (selected_class_code,),
                    )
                else:
                    return jsonify({"ok": True, "students": []}), 200
            elif class_code:
                cur.execute(
                    base_select + "WHERE s.class_code=%s ORDER BY s.id DESC",
                    (class_code.strip(),),
                )
            else:
                cur.execute(base_select + "ORDER BY s.id DESC")
            rows = cur.fetchall()
            student_ids = [int(r.get("id") or 0) for r in rows if int(r.get("id") or 0) > 0]
            subject_map: Dict[int, List[Dict[str, Any]]] = {}
            if student_ids:
                placeholders = ",".join(["%s"] * len(student_ids))
                cur.execute(
                    "SELECT ss.student_id, su.id AS subject_id, su.subject_code, su.subject_name "
                    "FROM student_subjects ss "
                    "JOIN subjects su ON su.id = ss.subject_id "
                    f"WHERE ss.student_id IN ({placeholders}) "
                    "ORDER BY su.subject_code ASC",
                    tuple(student_ids),
                )
                for srow in cur.fetchall() or []:
                    sid = int(srow.get("student_id") or 0)
                    subject_map.setdefault(sid, []).append(
                        {
                            "id": int(srow.get("subject_id") or 0),
                            "subject_code": srow.get("subject_code"),
                            "subject_name": srow.get("subject_name"),
                        }
                    )
        for r in rows:
            if isinstance(r.get("created_at"), datetime):
                r["created_at"] = r["created_at"].strftime("%Y-%m-%d %H:%M:%S")
            r["attended_today"] = bool(r.get("attended_today"))
            sid = int(r.get("id") or 0)
            subjects = subject_map.get(sid, [])
            r["subjects"] = subjects
            r["subject_ids"] = [int(x.get("id") or 0) for x in subjects]
        return jsonify({"ok": True, "students": rows}), 200
    finally:
        conn.close()


@app.route("/students", methods=["POST"])
def students_create():
    """
    JSON: student_code, full_name, email?, phone?, class_code?, notes?, face_images?: [base64,...]
    Ảnh được lưu vào data/<student_code>/ để train.
    """
    body = request.get_json(silent=True) or {}
    code = (body.get("student_code") or "").strip()
    name = (body.get("full_name") or "").strip()
    if not code or not name:
        return jsonify({"ok": False, "error": "student_code và full_name là bắt buộc."}), 400

    email = body.get("email")
    phone = body.get("phone")
    class_code = (body.get("class_code") or "").strip()
    notes = body.get("notes")
    face_images: List[str] = body.get("face_images") or []
    subject_ids = _parse_subject_ids(body.get("subject_ids"))

    if not class_code:
        return jsonify({"ok": False, "error": "Phải chọn lớp (class_code)."}), 400
    if not subject_ids:
        return jsonify({"ok": False, "error": "Phải chọn môn học (subject_ids)."}), 400

    if not face_images:
        _log_register(code, 0, "FAILED (missing_images)")
        return jsonify({"ok": False, "error": "Bắt buộc phải chụp ảnh mới để lưu sinh viên mới."}), 400

    # Chống trùng khuôn mặt khi đăng ký (KHÔNG phụ thuộc train/embeddings.npz)
    # So sánh trực tiếp với ảnh đã lưu trong data/<MSSV>/.
    if face_images:
        # A) Tạo embedding cho ảnh đăng ký (bỏ qua ảnh lỗi/không detect được mặt)
        new_embeddings: List[np.ndarray] = []
        last_err: Optional[str] = None
        for raw in face_images:
            if not raw:
                continue
            try:
                rgb = decode_base64_to_numpy_rgb(raw)
            except Exception:
                last_err = "decode_base64_to_numpy_rgb failed"
                continue
            try:
                face_crop_rgb = _detect_and_crop_face_rgb(rgb)
                if face_crop_rgb is not None:
                    face_rgb = cv2.resize(face_crop_rgb, _FACE_SIZE, interpolation=cv2.INTER_AREA)
                    emb = _embedding_from_face_rgb(face_rgb)
                else:
                    # fallback: để DeepFace tự detect mặt trên ảnh gốc
                    emb = embedding_from_image_rgb(rgb)
                new_embeddings.append(emb)
            except Exception as e:
                last_err = str(e)
                continue

        if len(new_embeddings) < 1:
            return (
                jsonify(
                    {
                        "ok": False,
                        "error": "Cần ít nhất 1 ảnh hợp lệ (nhìn rõ khuôn mặt) để đăng ký.",
                        "valid_faces": int(len(new_embeddings)),
                        "last_error": last_err,
                    }
                ),
                400,
            )

        # B) Lấy danh sách sinh viên hiện có để map student_code -> full_name
        code_to_name: Dict[str, str] = {}
        conn = db_connect()
        try:
            with conn.cursor() as cur:
                cur.execute("SELECT student_code, full_name FROM students")
                for r in cur.fetchall() or []:
                    sc = str(r.get("student_code") or "").strip()
                    fn = str(r.get("full_name") or "").strip()
                    if sc:
                        code_to_name[sc] = fn or sc
        finally:
            conn.close()

        # C) Build embeddings DB từ data/<MSSV>/ (tối đa 5–10 ảnh mỗi SV), có cache RAM
        def _read_image_rgb(path) -> Optional[np.ndarray]:
            try:
                buf = np.fromfile(str(path), dtype=np.uint8)
                bgr = cv2.imdecode(buf, cv2.IMREAD_COLOR)
                if bgr is None or bgr.size == 0:
                    return None
                return bgr[:, :, ::-1].copy()
            except Exception:
                return None

        max_images_per_student = int(os.environ.get("DUP_MAX_IMAGES_PER_STUDENT", "10"))
        max_faces_per_student = int(os.environ.get("DUP_MAX_FACES_PER_STUDENT", "6"))

        db_embeddings: List[np.ndarray] = []
        db_owners: List[str] = []

        data_root = data_dir()
        if data_root.is_dir():
            for stu_dir in sorted([p for p in data_root.iterdir() if p.is_dir()]):
                stu_code = stu_dir.name
                # skip thư mục trùng chính MSSV đang đăng ký (nếu tồn tại sẽ bị chặn ở bước sau)
                if stu_code == code:
                    continue

                imgs = []
                for ext in ("*.jpg", "*.jpeg", "*.png", "*.webp"):
                    imgs.extend(stu_dir.glob(ext))
                if not imgs:
                    continue
                # ổn định: sort theo mtime desc, lấy tối đa N ảnh
                imgs = sorted(imgs, key=lambda p: p.stat().st_mtime, reverse=True)[:max_images_per_student]
                sig = tuple((p.name, int(p.stat().st_mtime), int(p.stat().st_size)) for p in imgs)

                cached = _register_dup_cache.get("students", {}).get(stu_code) or {}
                cached_sig = cached.get("sig")
                cached_embs = cached.get("embeddings") if isinstance(cached.get("embeddings"), list) else None
                if cached_sig == sig and cached_embs:
                    embs = cached_embs
                else:
                    embs: List[np.ndarray] = []
                    for p in imgs:
                        rgb = _read_image_rgb(p)
                        if rgb is None:
                            continue
                        crop_rgb = _detect_and_crop_face_rgb(rgb)
                        if crop_rgb is None:
                            continue
                        try:
                            face_rgb = cv2.resize(crop_rgb, _FACE_SIZE, interpolation=cv2.INTER_AREA)
                            emb = _embedding_from_face_rgb(face_rgb)
                            embs.append(emb)
                        except Exception:
                            continue
                        if len(embs) >= max_faces_per_student:
                            break
                    _register_dup_cache.setdefault("students", {})[stu_code] = {"sig": sig, "embeddings": embs}

                if not embs:
                    continue
                owner = code_to_name.get(stu_code) or stu_code
                for emb in embs:
                    db_embeddings.append(emb)
                    db_owners.append(owner)

        # D) So sánh: max similarity trên toàn bộ (new_embeddings x db_embeddings)
        max_similarity = 0.0
        max_name = None
        if db_embeddings:
            for emb_new in new_embeddings:
                for emb_db, owner in zip(db_embeddings, db_owners):
                    sim = float(np.dot(emb_new, emb_db))
                    if sim > max_similarity:
                        max_similarity = sim
                        max_name = owner

        print("MAX SIMILARITY:", max_similarity)
        if max_similarity > 0.75:
            _log_register(code, int(len(face_images)), "DUPLICATE FACE DETECTED")
            return (
                jsonify(
                    {
                        "success": False,
                        "error": "Khuôn mặt đã tồn tại trong hệ thống",
                        "similarity": float(max_similarity),
                        "existing_student": max_name,
                    }
                ),
                409,
            )

    data_stu = data_dir() / code
    if data_stu.is_dir():
        _log_register(code, int(len(face_images or [])), "FAILED (data_folder_exists)")
        return jsonify(
            {
                "ok": False,
                "error": "Đã tồn tại thư mục ảnh trong data/ cho MSSV này. Xóa thư mục hoặc dùng MSSV khác.",
            }
        ), 409

    conn = db_connect()
    try:
        with conn.cursor() as cur:
            cur.execute("SELECT 1 FROM classes WHERE class_code=%s LIMIT 1", (class_code,))
            if not cur.fetchone():
                return jsonify({"ok": False, "error": "Lớp không tồn tại trong hệ thống."}), 400
            cur.execute("SELECT 1 FROM students WHERE student_code=%s LIMIT 1", (code,))
            if cur.fetchone():
                return jsonify({"ok": False, "error": "MSSV đã tồn tại."}), 409
            cur.execute(
                "INSERT INTO students (student_code, full_name, email, phone, class_code, notes) "
                "VALUES (%s,%s,%s,%s,%s,%s)",
                (code, name, email, phone, class_code, notes),
            )
            sid = cur.lastrowid
            _set_student_subjects(cur, int(sid), subject_ids)
    finally:
        conn.close()

    saved = 0
    if face_images:
        out_dir = data_dir() / code
        out_dir.mkdir(parents=True, exist_ok=True)
        for raw in face_images:
            try:
                rgb = decode_base64_to_numpy_rgb(raw)
            except Exception:
                continue
            from PIL import Image

            img = Image.fromarray(rgb)
            fname = f"{uuid.uuid4().hex[:10]}.jpg"
            img.save(str(out_dir / fname), quality=92)
            saved += 1

    if saved < 1:
        conn = db_connect()
        try:
            with conn.cursor() as cur:
                cur.execute("DELETE FROM students WHERE id=%s", (sid,))
        finally:
            conn.close()
        return jsonify({"ok": False, "error": "Không lưu được ảnh đăng ký. Vui lòng chụp lại."}), 400

    _log_register(code, int(saved), "SUCCESS")
    # JSON hợp lệ + Content-Type application/json (jsonify); client parse ổn định với status 200.
    return (
        jsonify(
            {
                "ok": True,
                "status": "success",
                "message": "Thêm sinh viên thành công",
                "id": sid,
                "student_code": code,
                "images_saved": saved,
            }
        ),
        200,
    )


@app.route("/students/<int:sid>", methods=["PUT"])
def students_update(sid: int):
    body = request.get_json(silent=True) or {}
    fields = []
    params: List[Any] = []
    subject_ids = _parse_subject_ids(body.get("subject_ids")) if "subject_ids" in body else None
    mapping = {
        "full_name": "full_name",
        "email": "email",
        "phone": "phone",
        "class_code": "class_code",
        "notes": "notes",
        "student_code": "student_code",
    }
    for k, col in mapping.items():
        if k in body:
            fields.append(f"{col}=%s")
            params.append(body[k])
    if not fields and subject_ids is None:
        return jsonify({"ok": False, "error": "Không có trường cập nhật."}), 400
    conn = db_connect()
    try:
        with conn.cursor() as cur:
            if fields:
                params.append(sid)
                sql = "UPDATE students SET " + ", ".join(fields) + " WHERE id=%s"
                cur.execute(sql, tuple(params))
                if cur.rowcount == 0:
                    return jsonify({"ok": False, "error": "Không tìm thấy sinh viên."}), 404
            else:
                cur.execute("SELECT id FROM students WHERE id=%s LIMIT 1", (sid,))
                if not cur.fetchone():
                    return jsonify({"ok": False, "error": "Không tìm thấy sinh viên."}), 404
            if subject_ids is not None:
                _set_student_subjects(cur, int(sid), subject_ids)
        return jsonify({"ok": True}), 200
    finally:
        conn.close()


@app.route("/students/<int:sid>", methods=["DELETE"])
def students_delete(sid: int):
    conn = db_connect()
    try:
        with conn.cursor() as cur:
            cur.execute("SELECT student_code FROM students WHERE id=%s", (sid,))
            row = cur.fetchone()
            if not row:
                return jsonify({"ok": False, "error": "Không tìm thấy sinh viên."}), 404
            code = row["student_code"]
            cur.execute("DELETE FROM students WHERE id=%s", (sid,))
        # Xoá ảnh train cục bộ (tuỳ chọn)
        import shutil

        folder = data_dir() / code
        if folder.is_dir():
            shutil.rmtree(folder, ignore_errors=True)
        return jsonify({"ok": True}), 200
    finally:
        conn.close()


@app.route("/attendance", methods=["POST"])
def attendance_create():
    """
    Lưu một lần điểm danh.
    JSON: student_code?, full_name?, class_code?, confidence?, status?,
          latitude?, longitude?, location_text?, student_db_id?
    """
    body = request.get_json(silent=True) or {}
    subject_id_raw = body.get("subject_id")
    try:
        subject_id = int(subject_id_raw) if subject_id_raw is not None and str(subject_id_raw).strip() != "" else None
    except Exception:
        subject_id = None

    # Nhánh mới: nếu gửi ảnh vào /attendance thì tự nhận diện (kèm liveness) rồi lưu.
    frame_rgb = _read_frame_rgb_from_request_json_or_multipart(body)
    if frame_rgb is not None:
        req_lat = _parse_request_gps_coordinate(body.get("latitude"))
        req_lng = _parse_request_gps_coordinate(body.get("longitude"))
        return _recognize_and_attend_from_frame_rgb(
            frame_rgb=frame_rgb,
            latitude=req_lat,
            longitude=req_lng,
            subject_id=subject_id,
        )

    status = (body.get("status") or "failed").strip()
    student_code = body.get("student_code")
    full_name = body.get("full_name")
    class_code = body.get("class_code")
    confidence = body.get("confidence")
    lat = body.get("latitude")
    lng = body.get("longitude")
    student_db_id_raw = body.get("student_db_id")

    if confidence is not None:
        try:
            confidence = float(confidence)
        except (TypeError, ValueError):
            confidence = None
    # Normalize: nếu client gửi confidence dạng % (0..100) thì đổi về 0..1
    # để khớp toàn bộ hệ thống (predict/recognize dùng 0..1).
    if confidence is not None and float(confidence) > 1.0:
        confidence = float(confidence) / 100.0

    # Không lưu lịch sử nếu điểm danh thất bại
    if str(status).strip().lower() != "success":
        return jsonify({"ok": True, "skipped": True}), 200
    # Ép ngưỡng: bắt buộc có confidence và phải >= RECOGNITION_THRESHOLD (mặc định 0.88) mới lưu.
    _min_rec = float(os.environ.get("RECOGNITION_THRESHOLD", "0.88"))
    if confidence is None:
        return (
            jsonify(
                {
                    "ok": True,
                    "success": False,
                    "skipped": True,
                    "message": "Điểm danh thất bại (thiếu confidence).",
                }
            ),
            200,
        )
    if float(confidence) < _min_rec:
        return (
            jsonify(
                {
                    "ok": True,
                    "success": False,
                    "skipped": True,
                    "message": "Điểm danh thất bại (confidence dưới ngưỡng nhận diện).",
                    "confidence": float(confidence),
                    "confidence_percent": round(float(confidence) * 100.0, 2),
                    "threshold": _min_rec,
                }
            ),
            200,
        )
    if not student_code and not body.get("student_db_id"):
        return jsonify({"ok": True, "skipped": True}), 200

    lat_p = _parse_request_gps_coordinate(lat)
    lng_p = _parse_request_gps_coordinate(lng)
    lat_u, lng_u, loc_text = _coords_for_attendance(lat_p, lng_p)

    student_db_id, student_code = _resolve_attendance_identity(student_db_id_raw, student_code)
    conn = None
    try:
        conn = db_connect()
        conn.autocommit(False)
        with conn.cursor() as cur:
            # Cùng logic với /predict: trùng theo (sinh viên + môn + ngày); subject_id NULL → fallback theo ngày.
            if _attendance_exists_for_student_subject_today(cur, student_db_id, student_code, subject_id):
                conn.rollback()
                return jsonify({"ok": True, "duplicate": True, "skipped": True}), 200
            cur.execute(
                "INSERT INTO attendance (student_db_id, student_code, full_name, class_code, confidence, status, "
                "latitude, longitude, location_text, subject_id) VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s,%s)",
                (student_db_id, student_code, full_name, class_code, confidence, status, lat_u, lng_u, loc_text, subject_id),
            )
            new_id = cur.lastrowid
        conn.commit()
        return jsonify({"ok": True, "id": new_id}), 201
    except Exception as e:
        if conn is not None:
            try:
                conn.rollback()
            except Exception:
                pass
        _log.exception("[ATTENDANCE] Lỗi khi lưu điểm danh (manual POST): %s", e)
        return jsonify(_json_fail("Lỗi khi lưu điểm danh", {"success": False})), 500
    finally:
        if conn is not None:
            try:
                conn.autocommit(True)
            except Exception:
                pass
            try:
                conn.close()
            except Exception:
                pass


@app.before_request
def _request_start_timer_and_log():
    # lưu thời gian để tính latency
    try:
        g._req_start = time.time()
    except Exception:
        g._req_start = None

    if not _should_log_requests():
        return

    # Có thể bật/tắt log health bằng env
    if request.path.rstrip("/") == "/health" and os.environ.get("LOG_HEALTH", "0") != "1":
        return

    _log.info(f"{request.method} {request.path} - START - {_client_ip()}")


@app.after_request
def _request_end_log(resp: WzResponse):
    if not _should_log_requests():
        return resp

    if request.path.rstrip("/") == "/health" and os.environ.get("LOG_HEALTH", "0") != "1":
        return resp

    start = getattr(g, "_req_start", None)
    ms = None
    if isinstance(start, (int, float)):
        ms = int((time.time() - float(start)) * 1000)

    # Log format theo yêu cầu: [time] METHOD PATH - status - ip
    ip = _client_ip()
    if ms is None:
        msg = f"{request.method} {request.path} - {resp.status_code} - {ip}"
    else:
        msg = f"{request.method} {request.path} - {resp.status_code} - {ip} - {ms}ms"

    # Nếu lỗi: log thêm reason + data (nếu có)
    if int(resp.status_code) >= 400:
        try:
            payload = resp.get_json(silent=True)
        except Exception:
            payload = None
        reason = None
        if isinstance(payload, dict):
            reason = payload.get("error") or payload.get("message")
        data = _safe_request_data_for_log()
        _log.error(
            "[ERROR] "
            + msg
            + (f"\nReason: {reason}" if reason else "")
            + (f"\nData: {_json_dumps_compact(data)}" if data else "")
        )
    else:
        _log.info(msg)

    return resp


@app.errorhandler(HTTPException)
def _log_http_exception(e: HTTPException):
    # Không thay đổi body/format response, chỉ log.
    try:
        ip = _client_ip()
        data = _safe_request_data_for_log()
        _log.error(
            f"[ERROR] {request.path} - {e.code} - {ip}\nReason: {e.name}\nData: {_json_dumps_compact(data)}"
        )
    except Exception:
        pass
    return e


@app.errorhandler(Exception)
def _log_unhandled_exception(e: Exception):
    # Lỗi server 500: log traceback + request data
    try:
        ip = _client_ip()
        data = _safe_request_data_for_log()
        tb = "".join(traceback.format_exception(type(e), e, e.__traceback__))
        _log.error(
            f"[ERROR] {request.method} {request.path} - 500 - {ip}\nReason: {type(e).__name__}: {e}\n"
            f"Data: {_json_dumps_compact(data)}\nTraceback:\n{tb}"
        )
    except Exception:
        # tránh vòng lặp log lỗi
        pass
    return jsonify({"ok": False, "error": "Internal Server Error"}), 500


@app.route("/attendance", methods=["GET"])
def attendance_list():
    """
    Query: class_code?/ma_lop?, from_date?, to_date?, status?, limit?
    """
    class_code = request.args.get("class_code")
    # alias theo spec UI
    ma_lop = request.args.get("ma_lop")
    from_date = request.args.get("from_date")
    to_date = request.args.get("to_date")
    status = request.args.get("status")
    limit = min(int(request.args.get("limit", "200")), 500)

    clauses: List[str] = []
    params: List[Any] = []
    class_filter = (ma_lop or class_code)
    if class_filter:
        # lọc theo lớp dựa trên class_code lưu trên students/attendance
        clauses.append("COALESCE(s.class_code, a.class_code) = %s")
        params.append(class_filter.strip())
    if from_date:
        clauses.append("DATE(a.checked_at) >= %s")
        params.append(from_date)
    if to_date:
        clauses.append("DATE(a.checked_at) <= %s")
        params.append(to_date)
    if status:
        clauses.append("a.status = %s")
        params.append(status)
    # build WHERE cho subquery (alias a0/s0) để dedupe
    clauses0: List[str] = []
    for cl in clauses:
        clauses0.append(cl.replace("a.", "a0.").replace("s.", "s0."))
    where0 = ("WHERE " + " AND ".join(clauses0)) if clauses0 else ""

    # Deduplicate: mỗi sinh viên (student_code hoặc student_db_id) mỗi ngày lấy 1 bản ghi mới nhất.
    # Dùng GROUP BY + MAX(id) để tương thích MySQL không có window function.
    sql = (
        "SELECT "
        "COALESCE(s.full_name, a.full_name) AS name, "
        "COALESCE(s.student_code, a.student_code) AS mssv, "
        "COALESCE(c.class_name, s.class_code, a.class_code) AS class_name, "
        "sub.subject_name AS subject_name, "
        "a.checked_at AS checked_at, "
        "a.status AS status, "
        "a.latitude AS latitude, "
        "a.longitude AS longitude, "
        "a.location_text AS location_text, "
        "a.id AS id, "
        "a.student_db_id AS student_db_id, "
        "a.student_code AS student_code, "
        "a.full_name AS full_name, "
        "a.class_code AS class_code, "
        "a.confidence AS confidence "
        "FROM attendance a "
        "JOIN ("
        "  SELECT "
        "    COALESCE(a0.student_code, CONCAT('db:', COALESCE(a0.student_db_id, 0))) AS ident, "
        # NOTE: dùng %% để tránh PyMySQL hiểu nhầm %Y/%m...
        "    DATE(a0.checked_at) AS day_key, "
        "    MAX(a0.id) AS max_id "
        "  FROM attendance a0 "
        "  JOIN students s0 ON (a0.student_db_id = s0.id OR a0.student_code = s0.student_code) "
        f"  {where0} "
        "  GROUP BY ident, day_key "
        ") t ON a.id = t.max_id "
        "JOIN students s ON (a.student_db_id = s.id OR a.student_code = s.student_code) "
        "LEFT JOIN classes c ON (COALESCE(s.class_code, a.class_code) = c.class_code) "
        "LEFT JOIN subjects sub ON (sub.id = a.subject_id) "
        "ORDER BY a.checked_at DESC, a.id DESC "
        "LIMIT %s"
    )
    params.append(limit)

    conn = db_connect()
    try:
        with conn.cursor() as cur:
            cur.execute(sql, tuple(params))
            rows = cur.fetchall()
        items = []
        for r in rows:
            out: Dict[str, Any] = dict(r or {})
            for k, v in list(out.items()):
                if isinstance(v, datetime):
                    out[k] = v.strftime("%Y-%m-%d %H:%M:%S")
                elif isinstance(v, Decimal):
                    out[k] = float(v)

            # map response theo spec UI (không ghép trùng lat,lng + location_text cùng tọa độ)
            location = _format_attendance_location_display(
                out.get("latitude"),
                out.get("longitude"),
                out.get("location_text"),
            )

            mapped = {
                "name": out.get("name"),
                "mssv": out.get("mssv"),
                "class": out.get("class_name"),
                "subject_name": out.get("subject_name"),
                "time": out.get("checked_at"),
                "status": out.get("status"),
                "location": location,
                # Luôn trả tọa độ rõ ràng cho UI (kể cả 0.0) — tránh client hiểu nhầm null/thiếu field.
                "latitude": out.get("latitude"),
                "longitude": out.get("longitude"),
                "location_text": out.get("location_text"),
            }
            out.update(mapped)
            items.append(out)

        # trả song song `items` (cũ) và `data` (mới) để không làm hỏng client khác
        return jsonify({"ok": True, "items": items, "data": items}), 200
    finally:
        conn.close()


@app.route("/export-excel", methods=["GET"])
def export_excel():
    """
    Xuất Excel lịch sử điểm danh từ database.
    """
    try:
        from openpyxl import Workbook
        from openpyxl.utils import get_column_letter
    except Exception as e:
        return (
            jsonify(
                {
                    "ok": False,
                    "error": f"Thiếu thư viện openpyxl ({e}). Cài bằng: pip install openpyxl",
                }
            ),
            500,
        )

    ma_lop = (request.args.get("ma_lop") or "").strip()
    class_code = (request.args.get("class_code") or "").strip()
    date_str = (request.args.get("date") or "").strip()

    clauses: List[str] = []
    params: List[Any] = []
    cls = ma_lop or class_code
    if cls:
        clauses.append("COALESCE(s.class_code, a.class_code) = %s")
        params.append(cls)
    if date_str:
        clauses.append("DATE(a.checked_at) = %s")
        params.append(date_str)
    where = ("WHERE " + " AND ".join(clauses)) if clauses else ""

    conn = db_connect()
    try:
        with conn.cursor() as cur:
            cur.execute(
                "SELECT "
                "COALESCE(s.full_name, a.full_name) AS full_name, "
                "COALESCE(s.student_code, a.student_code) AS student_code, "
                "COALESCE(s.class_code, a.class_code) AS class_code, "
                "a.checked_at AS checked_at, "
                "a.status AS status, "
                "a.confidence AS confidence, "
                "a.location_text AS location_text "
                "FROM attendance a "
                "LEFT JOIN students s ON (a.student_db_id = s.id) "
                f"{where} "
                "ORDER BY a.checked_at DESC, a.id DESC",
                tuple(params),
            )
            rows = cur.fetchall() or []
    finally:
        conn.close()

    wb = Workbook()
    ws = wb.active
    ws.title = "Attendance"

    ws.append(
        [
            "Tên sinh viên",
            "MSSV",
            "Lớp",
            "Thời gian",
            "Trạng thái",
            "Độ tin cậy",
            "Vị trí",
        ]
    )

    def _fmt_dt(v: Any) -> str:
        if isinstance(v, datetime):
            return v.strftime("%d/%m/%Y %H:%M:%S")
        return str(v) if v is not None else ""

    for r in rows:
        ws.append(
            [
                r.get("full_name") or "",
                r.get("student_code") or "",
                r.get("class_code") or "",
                _fmt_dt(r.get("checked_at")),
                r.get("status") or "",
                float(r.get("confidence") or 0),
                r.get("location_text") or "",
            ]
        )

    # Auto width cột đơn giản (dựa trên độ dài text)
    for col_idx in range(1, ws.max_column + 1):
        col_letter = get_column_letter(col_idx)
        max_len = 0
        for cell in ws[col_letter]:
            try:
                max_len = max(max_len, len(str(cell.value or "")))
            except Exception:
                pass
        ws.column_dimensions[col_letter].width = min(48, max(12, max_len + 2))

    file_stream = io.BytesIO()
    wb.save(file_stream)
    file_stream.seek(0)

    return send_file(
        file_stream,
        as_attachment=True,
        download_name="attendance.xlsx",
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )


_initialized = False


@app.before_request
def _bootstrap_once():
    """Khởi tạo schema + preload model một lần khi process nhận request đầu tiên."""
    global _initialized
    if request.path.rstrip("/") == "/health":
        return
    if _initialized:
        return
    init_schema()
    try:
        get_yolo_detector()
    except Exception as e:
        _log.warning(f"[BOOT] Không preload được YOLO: {e}")
    _initialized = True


@app.before_request
def _log_incoming_request():
    if not _should_log_requests():
        return
    # tránh spam quá nhiều từ health checks nếu muốn
    if request.path.rstrip("/") == "/health" and os.environ.get("LOG_HEALTH", "0") != "1":
        return
    try:
        print(f"[REQ] {request.method} {request.path}")
    except Exception:
        pass


@app.after_request
def _log_outgoing_response(resp: WzResponse):
    if not _should_log_requests():
        return resp
    if request.path.rstrip("/") == "/health" and os.environ.get("LOG_HEALTH", "0") != "1":
        return resp
    try:
        print(f"[RES] {request.method} {request.path} -> {resp.status_code}")
    except Exception:
        pass
    return resp


@app.before_request
def _require_auth():
    if request.method == "OPTIONS":
        return
    path = request.path.rstrip("/") or "/"
    if path == "/health" or path == "/auth/login":
        return
    auth = request.headers.get("Authorization", "")
    if not auth.startswith("Bearer "):
        return jsonify({"ok": False, "error": "Unauthorized"}), 401
    payload = verify_auth_token(auth[7:].strip())
    if not payload:
        return jsonify({"ok": False, "error": "Token không hợp lệ hoặc đã hết hạn."}), 401
    g.admin = payload


if __name__ == "__main__":
    # Preload YOLO ngay lúc start để log hiện trực tiếp trên CMD.
    try:
        get_yolo_detector()
    except Exception as e:
        _log.warning(f"[BOOT] YOLO preload failed at startup: {e}")
    port = int(os.environ.get("PORT", "5000"))
    app.run(host="0.0.0.0", port=port, debug=os.environ.get("FLASK_DEBUG") == "1")
